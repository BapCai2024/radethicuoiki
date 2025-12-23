\
from __future__ import annotations
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional
import openpyxl

from .utils import safe_int, round_to_step

@dataclass
class LessonRow:
    tt: int
    topic: str
    lesson: str
    periods: int
    ratio_pct: float
    points_target: float
    counts: Dict[Tuple[str,int], int]  # (qtype, level) -> count

@dataclass
class MatrixTemplate:
    title: str
    grade: Optional[int]
    subject: Optional[str]
    semester: Optional[str]
    lessons: List[LessonRow]
    points_per_qtype: Dict[str, float]  # qtype -> points
    total_points: float

QTYPE_COLS = {
    "MCQ": ("G","H","I"),
    "TF": ("J","K","L"),
    "MATCH": ("M","N","O"),
    "FILL": ("P","Q","R"),
    "ESSAY": ("S","T","U"),
}

QTYPE_LABELS_VI = {
    "MCQ": "Nhiều lựa chọn",
    "TF": "Đúng-Sai",
    "MATCH": "Nối cột",
    "FILL": "Điền khuyết",
    "ESSAY": "Tự luận",
}

def load_matrix_template(xlsx_path: str, total_points: float = 10.0, step: float = 0.25) -> MatrixTemplate:
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb["ma trận"] if "ma trận" in wb.sheetnames else wb[wb.sheetnames[0]]

    title = str(ws["C2"].value or "MA TRẬN").strip()
    # Heuristic extract grade/subject/semester from title
    grade = None
    subject = None
    semester = None
    t_up = title.upper()
    for g in range(1,6):
        if f" {g} " in f" {t_up} " or f"LỚP {g}" in t_up:
            grade = g
    if "TIN" in t_up:
        subject = "Tin"
    if "HỌC KÌ I" in t_up or "HKI" in t_up or "HK1" in t_up:
        semester = "HK1"
    if "HỌC KÌ II" in t_up or "HKII" in t_up or "HK2" in t_up:
        semester = "HK2"

    # Find lesson rows: start after header (row 6), until a row where col A == 'Tổng số câu'
    start_row = 7
    end_row = None
    for r in range(start_row, ws.max_row+1):
        a = ws.cell(r, 1).value
        if isinstance(a, str) and "Tổng số câu" in a:
            end_row = r-1
            break
    if end_row is None:
        end_row = ws.max_row

    # Total periods
    total_periods = 0
    for r in range(start_row, end_row+1):
        total_periods += safe_int(ws.cell(r, 4).value, 0)

    # Default points per qtype from rows 22-26 col D in provided template; fallback to step values.
    points_per_qtype = {"MCQ": 0.25, "TF": 0.25, "MATCH": 0.5, "FILL": 0.25, "ESSAY": 0.5}
    # Attempt read by scanning for "Điểm 1 câu" in column C
    for r in range(end_row+1, ws.max_row+1):
        c_val = str(ws.cell(r, 3).value or "")
        d_val = ws.cell(r, 4).value
        if "Điểm 1 câu" in c_val and d_val is not None:
            v = round_to_step(float(d_val), step)
            if "nhiều" in c_val.lower():
                points_per_qtype["MCQ"] = v
            elif "đúng" in c_val.lower():
                points_per_qtype["TF"] = v
            elif "nối" in c_val.lower():
                points_per_qtype["MATCH"] = v
            elif "điền" in c_val.lower():
                points_per_qtype["FILL"] = v
            elif "tự luận" in c_val.lower():
                points_per_qtype["ESSAY"] = v

    lessons: List[LessonRow] = []
    current_topic = ""
    for r in range(start_row, end_row+1):
        tt = ws.cell(r, 1).value
        if tt is None:
            continue
        try:
            tt_int = int(tt)
        except Exception:
            continue
        topic = ws.cell(r, 2).value
        if topic:
            current_topic = str(topic).strip()
        lesson = str(ws.cell(r, 3).value or "").strip()
        periods = safe_int(ws.cell(r, 4).value, 0)
        ratio = (periods / total_periods * 100.0) if total_periods else 0.0
        points_target = total_points * ratio / 100.0

        counts: Dict[Tuple[str,int], int] = {}
        for qtype, cols in QTYPE_COLS.items():
            for level, col_letter in zip([1,2,3], cols):
                cell_val = ws[f"{col_letter}{r}"].value
                counts[(qtype, level)] = safe_int(cell_val, 0)

        lessons.append(LessonRow(
            tt=tt_int,
            topic=current_topic,
            lesson=lesson,
            periods=periods,
            ratio_pct=ratio,
            points_target=points_target,
            counts=counts
        ))

    return MatrixTemplate(
        title=title,
        grade=grade,
        subject=subject,
        semester=semester,
        lessons=lessons,
        points_per_qtype=points_per_qtype,
        total_points=total_points
    )
